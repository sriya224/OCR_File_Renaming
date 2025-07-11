import os
import re
import json
import argparse
import pandas as pd
from collections import defaultdict, Counter

def parse_ocr_file(filepath):
    with open(filepath, "r", encoding="utf-8") as f:
        lines = f.read().splitlines()

    grouped = defaultdict(list)
    current_file = None
    header_pattern = re.compile(r"--- (.+?) ---")

    for line in lines:
        match = header_pattern.match(line.strip())
        if match:
            current_file = match.group(1).replace("_label.png", "")
        elif current_file and line.strip():
            grouped[current_file].append(line.strip())

    return grouped

def extract_study_id_and_clean(lines):
    study_id_pattern = re.compile(r"\b\d{2}-\d{3}\b")
    date_pattern = re.compile(r"\d{1,2}/\d{1,2}/\d{2,4}")
    time_pattern = re.compile(r"\d{1,2}:\d{2} ?(AM|PM)?", re.IGNORECASE)
    noise_terms = {"glintlab"}

    study_id = None
    filename_tokens = []

    for line in lines:
        if date_pattern.search(line) or time_pattern.search(line):
            continue

        words = line.strip().split()
        for word in words:
            cleaned = word.strip(".,;:*+").replace("/", "-")
            if not cleaned or cleaned.lower() in noise_terms:
                continue

            if study_id is None and study_id_pattern.match(cleaned):
                study_id = cleaned

            filename_tokens.append(cleaned)

            if study_id_pattern.match(cleaned):
                break

    filename = "_".join(filename_tokens)
    return study_id, filename

def rename_mrxs_files(ocr_txt_path, mrxs_folder, label_folder):
    grouped = parse_ocr_file(ocr_txt_path)

    rename_map = {}
    study_id_to_files = defaultdict(list)
    records = []  # For Excel output

    for fname, lines in grouped.items():
        study_id, cleaned_name = extract_study_id_and_clean(lines)
        rename_map[fname] = cleaned_name
        if study_id:
            study_id_to_files[study_id].append(fname)

        label_img = f"{fname}_label.png"
        records.append({
            "Original File Name": fname + ".mrxs",
            "Label Image": label_img,
            "New File Name": cleaned_name + ".mrxs"
        })
    export_excel_with_images(records, label_folder)


    renamed = []
    for old_name, new_name in rename_map.items():
        old_path = os.path.join(mrxs_folder, old_name + "_label.png")
        new_path = os.path.join(mrxs_folder, new_name + "_label_1.png")
        if os.path.exists(old_path):
            if not os.path.exists(new_path):
                os.rename(old_path, new_path)
                renamed.append((old_name, new_name))
            else:
                print(f"Skipped: {new_name}.mrxs already exists.")
        else:
            print(f"File not found: {old_path}")

    print(f"\nRenamed {len(renamed)} file(s).")
    for old, new in renamed:
        print(f" - {old}.mrxs → {new}.mrxs")

    with open("renamed_files.json", "w", encoding="utf-8") as f:
        json.dump(rename_map, f, indent=2)
    print("Saved mapping to renamed_files.json")

    # Outlier detection
    print("\n🔍 Checking for outlier study IDs...")
    outliers = {sid: files for sid, files in study_id_to_files.items() if len(files) == 1}
    if outliers:
        print(f"\n⚠️ Found {len(outliers)} potential outlier study ID(s):")
        for sid, files in outliers.items():
            print(f" - Study ID '{sid}' only found in file: {files[0]}")
    else:
        print("✅ No study ID outliers detected.")

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

def export_excel_with_images(records, label_folder, output_path="filename_mapping_with_images.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Renamed Files"

    # Write header
    ws.append(["Label Image", "Original File Name", "New File Name"])

    for i, record in enumerate(records, start=2):  # Excel rows start at 1; row 1 is header
        label_img_path = os.path.join(label_folder, record["Label Image"])
        if os.path.exists(label_img_path):
            # Resize image to thumbnail
            thumb_path = f"{label_img_path}_thumb.png"
            with PILImage.open(label_img_path) as img:
                img.thumbnail((200, 200))
                img.save(thumb_path)

            img_for_excel = XLImage(thumb_path)
            img_for_excel.anchor = f"A{i}"
            ws.add_image(img_for_excel)

        # Write file names
        ws.cell(row=i, column=2, value=record["Original File Name"])
        ws.cell(row=i, column=3, value=record["New File Name"])

        # Resize row hieght
        ws.row_dimensions[i].height = 120

    # Resize column width
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30

    wb.save(output_path)
    print(f"\n📄 Saved Excel mapping with images to {output_path}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Rename .mrxs files using OCR results and export Excel mapping.")
    parser.add_argument("--ocr", required=True, help="Path to vision_ocr_results.txt")
    parser.add_argument("--folder", required=True, help="Path to folder containing .mrxs files")
    parser.add_argument("--labels", required=True, help="Path to folder containing label .png images")
    args = parser.parse_args()

    rename_mrxs_files(args.ocr, args.folder, args.labels)
