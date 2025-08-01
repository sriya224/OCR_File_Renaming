
import os
import re
import argparse
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

def log(message):
    print(message)
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(message + "\n")

def parse_ocr_file(filepath):
    with open(filepath, "r", encoding="utf-8") as f:
        lines = f.read().splitlines()

    grouped = defaultdict(list)
    confidences = defaultdict(list)
    current_file = None
    header_pattern = re.compile(r"--- (.+?) ---")

    for line in lines:
        match = header_pattern.match(line.strip())
        if match:
            current_file = match.group(1).replace(".png", "")
            grouped[current_file] = []
        elif current_file and line.strip():
            grouped[current_file].append(line.strip())
            if line.startswith("[") and "]" in line:
                try:
                    conf = float(line.split("]")[0].strip("["))
                    confidences[current_file].append(conf)
                except:
                    continue
    return grouped, confidences

def extract_study_id_and_clean(lines):
    study_id_pattern = re.compile(r"\b\d{2}-\d{3}\b")
    noise_terms = {"glintlab"}
    study_id = None
    filename_tokens = []

    for line in lines:
        words = line.strip().split()
        for word in words:
            cleaned = word.strip(".,;:*+").replace("/", "-")
            if not cleaned or cleaned.lower() in noise_terms:
                continue
            if study_id is None and study_id_pattern.match(cleaned):
                study_id = cleaned
            filename_tokens.append(cleaned)
            if study_id_pattern.match(cleaned):
                break
    filename = "_".join(filename_tokens)
    return study_id, filename

def export_excel_with_images(records, label_folder, output_folder, output_path="expanded_ocr_results.xlsx", threshold=0.693):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "OCR Mapping"

    headers = ["Label Image", "Avg Confidence", "Original File Name", "New File Name"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True, size=12)

    for i, record in enumerate(records, start=2):
        label_img_path = os.path.join(label_folder, record["Label Image"])
        if os.path.exists(label_img_path):
            thumb_path = f"{label_img_path}_thumb.png"
            with PILImage.open(label_img_path) as img:
                img.thumbnail((200, 200))
                img.save(thumb_path)
            img_for_excel = XLImage(thumb_path)
            img_for_excel.anchor = f"A{i}"
            ws.add_image(img_for_excel)

        conf_cell = ws.cell(row=i, column=2, value=record["Confidence"])
        if record["Confidence"] < threshold:
            conf_cell.font = Font(color="FF0000", bold=True)

        ws.cell(row=i, column=3, value=record["Original File Name"])
        ws.cell(row=i, column=4, value=record["New File Name"])
        ws.row_dimensions[i].height = 120

    for col in ["A", "B", "C", "D"]:
        ws.column_dimensions[col].width = 35

    path = os.path.join(output_folder, output_path)
    wb.save(path)
    log(f"\nSaved Excel mapping with images and confidence to {output_path}")

def postprocess_ocr(ocr_txt_path, mrxs_folder, output_folder, label_folder):
    global LOG_PATH
    LOG_PATH = os.path.join(output_folder, "rename_log.txt")
    log(f"OCR input: {ocr_txt_path}")
    log(f"Output folder: {output_folder}")

    grouped, confidence_scores = parse_ocr_file(ocr_txt_path)
    study_id_to_files = defaultdict(list)
    records = []

    for fname, lines in grouped.items():
        study_id, cleaned_name = extract_study_id_and_clean(lines)
        avg_conf = round(sum(confidence_scores[fname]) / len(confidence_scores[fname]), 3) if confidence_scores[fname] else 0
        if study_id:
            study_id_to_files[study_id].append(fname)
        label_img = f"{fname}.png"
        records.append({
            "Original File Name": fname + ".mrxs",
            "Label Image": label_img,
            "New File Name": cleaned_name + ".mrxs",
            "Confidence": avg_conf
        })

    export_excel_with_images(records, label_folder, output_folder)

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Post-process OCR results into Excel mapping without renaming.")
    parser.add_argument("--ocr", required=True, help="Path to vision_ocr_results.txt")
    parser.add_argument("--folder", required=True, help="Path to folder containing .mrxs files")
    parser.add_argument("--labels", required=True, help="Path to folder containing label .png images")
    parser.add_argument("-o", required=True, help="Path to store output files")
    args = parser.parse_args()

    postprocess_ocr(args.ocr, args.folder, args.o, args.labels)
